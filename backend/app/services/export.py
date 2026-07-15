import csv
import io
import logging
from datetime import datetime, timezone

from app.config import settings

logger = logging.getLogger(__name__)


def export_to_csv(items: list[dict]) -> bytes:
    """Export shopping list items to CSV format."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["分类", "商品名称", "品牌", "规格", "数量", "单位", "优选渠道", "到手价", "链接", "备注"])
    for item in items:
        writer.writerow([
            item.get("category_name", ""),
            item.get("product_name", ""),
            item.get("brand", ""),
            item.get("specification", ""),
            item.get("quantity", 1),
            item.get("unit", "个"),
            item.get("best_channel", ""),
            item.get("best_price", ""),
            item.get("product_url", ""),
            item.get("notes", ""),
        ])
    return output.getvalue().encode("utf-8-sig")


def export_to_excel(items: list[dict]) -> bytes:
    """Export shopping list items to Excel format using openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        logger.warning("openpyxl not installed, falling back to CSV")
        return export_to_csv(items)

    wb = Workbook()
    ws = wb.active
    ws.title = "采购清单"

    headers = ["分类", "商品名称", "品牌", "规格", "数量", "单位", "优选渠道", "到手价", "链接", "备注"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for row_idx, item in enumerate(items, 2):
        values = [
            item.get("category_name", ""),
            item.get("product_name", ""),
            item.get("brand", ""),
            item.get("specification", ""),
            item.get("quantity", 1),
            item.get("unit", "个"),
            item.get("best_channel", ""),
            item.get("best_price", ""),
            item.get("product_url", ""),
            item.get("notes", ""),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_to_pdf(items: list[dict]) -> bytes:
    """Export shopping list items to PDF."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        logger.warning("reportlab not installed, falling back to CSV")
        return export_to_csv(items)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()

    elements = []
    elements.append(Paragraph("家庭采购清单", styles["Title"]))
    elements.append(Spacer(1, 20))

    headers = ["分类", "商品名称", "数量", "优选渠道", "到手价", "备注"]
    data = [headers]
    for item in items:
        data.append([
            item.get("category_name", ""),
            item.get("product_name", ""),
            str(item.get("quantity", 1)),
            item.get("best_channel", ""),
            str(item.get("best_price", "")),
            item.get("notes", ""),
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)

    doc.build(elements)
    return buffer.getvalue()
